from .models import *
import openpyxl
import pandas as pd
from openpyxl import Workbook
import collections
from math import isnan
from django.shortcuts import get_object_or_404
from datetime import date
tableNames = ['BIOCHEM-01','BIOCHEM-02','BIOCHEM-03','BIOCHEM-04','BIOCHEM-05',
				'BIOCHEM-06','BIOCHEM-07','BIOCHEM-08','CBA-01','CBA-02','CBA-03','ENDO-01','FC-01','FC-03','FC-04',
				'FC-07','FC-08','HEM-01','HPIBD-01','HPIBD-02','HPIBD-03','HPIBD-04','HPA-02','HPNI-01','IINFLC-01',
				'IINFLC-02','IINFLC-03','IINFLC-04','IINFLC-05','IINFLC-06','NI-01','NI-02-GRS-01','NI-02-OFD-01',
				'NI-02-ROT-01','PR-02','AR-02','AR-03','AR-04','AR-05','AR-06','AR-07','GI-05','GI-06',"info"
				]

def create_mouseHash(data):
	for i in range(len(data["Mouse ID"])):
		m = Mouse()
		#print("MOUSE")

		for key in data:
			if ('Mouse ID' in key):
				m.mid = data[key][i]			
			elif ('genotype' in key.lower()):
				m.genotype = data[key][i]
			elif ('strain' in key.lower()):
				m.strain = data[key][i]
			elif ('tail' in key.lower()):
				m.tail_num = data[key][i]				 
			elif ('induced' in key.lower()):
				m.induced = data[key][i]
			elif ('treated' in key.lower() or 'treatement' in key.lower()):
				m.treated = data[key][i]
			elif ('gender' in key.lower() or 'sex' in key.lower()):
				m.gender = data[key][i] 
			elif ('Age' in key):
				m.age = data[key][i] 
			elif ('birth' in key.lower()):
				m.dateofBirth = data[key][i] 				
			elif ('diet' in key.lower()):
				m.diet = data[key][i] 
			elif ('cage' in key.lower()):
				m.mouse_info = data[key][i]
			elif ('environmental' in key.lower()):
				m.other = data[key][i]
			elif ('report' in key.lower()):
				m.health_report = data[key][i]
		if(Mouse.objects.filter(mid=m.mid, genotype=m.genotype,dateofBirth=m.dateofBirth)):
			print('mphke edw {}',m.mid)
			continue;
		print("Saving mouse {}".format(m.mid)) 	
		m.save()

def data_iinflc04(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		iinflc04 = Iinflc04()
		iinflc04.assayid = assay
		for key in data:
			if(flag):
				if ('ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						iinflc04.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					iinflc04.timepoint= data[key][i]
				elif ('age' in key.lower()):
					iinflc04.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					iinflc04.measurement_date= d				 
				elif ('weight' in key.lower()):
					iinflc04.weight= data[key][i]
				elif ('comment' in key.lower()):
					iinflc04.comment= data[key][i]
		if(flag):				
			iinflc04.save()

def data_iinflc01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Iinflc01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('weight' in key.lower()):
					if ('score' in key.lower()):
						experiment.weight_score = data[key][i]
					else:
						experiment.weight = data[key][i]
				elif ('stool' in key.lower()):
					if ('consistency' in key.lower()):
						experiment.stool_consistency = data[key][i]
					else:
						experiment.blood_in_stool = data[key][i]
				elif ('dai' in key.lower()):
					experiment.dai= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_iinflc02(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Iinflc02()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sample source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('total cell' in key.lower()):
					if ('small' in key.lower()):
						experiment.cell_count_s_intestine = data[key][i]
					else:
						experiment.cell_count_l_intestine = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_iinflc03(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Iinflc03()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sum intensity' in key.lower()):
					experiment.sum_intensity= data[key][i]
				elif ('net intensity' in key.lower()):
					experiment.net_intensity = data[key][i]
				elif ('mean intensity' in key.lower()):
					experiment.mean_intensity = data[key][i]
				elif ('acquisition' in key.lower()):
					experiment.acquisition_settings = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_iinflc05(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Iinflc05()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sum intensity' in key.lower()):
					experiment.sum_intensity= data[key][i]
				elif ('net intensity' in key.lower()):
					experiment.net_intensity = data[key][i]
				elif ('mean intensity' in key.lower()):
					experiment.mean_intensity = data[key][i]
				elif ('acquisition' in key.lower()):
					experiment.acquisition_settings = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_iinflc06(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Iinflc06()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sum intensity' in key.lower()):
					experiment.sum_intensity= data[key][i]
				elif ('net intensity' in key.lower()):
					experiment.net_intensity = data[key][i]
				elif ('mean intensity' in key.lower()):
					experiment.mean_intensity = data[key][i]
				elif ('acquisition' in key.lower()):
					experiment.acquisition_settings = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_ni01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		ni01 = Ni01()
		ni01.assayid = assay
		for key in data:
			if(flag):
				if ('ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						ni01.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					ni01.timepoint= data[key][i]
				elif ('age' in key.lower()):
					ni01.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					ni01.measurement_date= d				 
				elif ('clinical' in key.lower()):
					ni01.clinical_score= data[key][i]
				elif ('comment' in key.lower()):
					ni01.comment= data[key][i]
		if(flag):					
			ni01.save()

def data_ni02rot01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Ni02Rot01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					experiment.timepoint= data[key][i]
				elif ('age' in key.lower()):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('individual latency' in key.lower()):
					if('1' in key.lower()):
						experiment.individual_latency_fall1 = data[key][i]
					else:
						experiment.individual_latency_fall2 = data[key][i]
				elif ('speed' in key.lower()):
					if('1' in key.lower()):
						experiment.speed_fall1 = data[key][i]
					else:
						experiment.speed_fall2 = data[key][i]
				elif ('mean latency' in key.lower()):
					experiment.mean_latency_fall= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]				
		if(flag):				
			experiment.save()

def data_ni02ofd01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Ni02ofd01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					experiment.timepoint= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('distance' in key.lower()):
					if('arena' in key.lower()):
						experiment.total_distance_wa = data[key][i]
					elif('peripheral' in key.lower()): 
						experiment.total_distance_pz = data[key][i]
					elif('central' in key.lower()):
						experiment.total_distance_cz = data[key][i]
				elif ('time spent' in key.lower()):
					if('peripheral' in key.lower()):
						experiment.time_pz = data[key][i]
					elif('central' in key.lower()):
						experiment.time_cz = data[key][i]
				elif ('speed' in key.lower()):
					if('arena' in key.lower()):
						experiment.avg_speed_wa = data[key][i]
					elif('peripheral' in key.lower()): 
						experiment.avg_speed_pz = data[key][i]
					elif('central' in key.lower()):
						experiment.avg_speed_cz = data[key][i]
		if(flag):				
			experiment.save()

def data_ni02grs01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Ni02grs01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					experiment.timepoint= data[key][i]
				elif ('measurement date' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('body weight (g)' in key.lower()):
					experiment.weight= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				elif ('forelimb grip' in key.lower()):
					#print('forelimb grip: ',data[key][i])
					if('measurement 1' in key.lower()):
						experiment.forelimb_r1 = data[key][i]
					elif('measurement 2' in key.lower()):
						experiment.forelimb_r2 = data[key][i]
					elif('measurement 3' in key.lower()):
						experiment.forelimb_r3 = data[key][i]
					elif('measurement mean' in key.lower()):
						#print('pass 4: ',data[key][i])						
						experiment.forelimb = data[key][i]
					elif('ratio' in key.lower()):
						experiment.forelimb_mean_ratio = data[key][i]
				elif ('hindlimb' in key.lower()):
					if('measurement 1' in key.lower()):
						experiment.hindlimb_r1 = data[key][i]
					elif('measurement 2' in key.lower()): 
						experiment.hindlimb_r2 = data[key][i]
					elif('measurement 3' in key.lower()):
						experiment.hindlimb_r3 = data[key][i]
					elif('measurement mean' in key.lower()):
						experiment.hindlimb = data[key][i]
					else:
						#print('pass 5: ',data[key][i])
						experiment.hindlimb_mean_ratio = data[key][i]	
		if(flag):				
			experiment.save()

def data_hem01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Hem01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' == key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					experiment.timepoint= data[key][i]
				elif ('measurement date' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				elif ('sample id' in key.lower()):
					experiment.sample_id= data[key][i]
				elif ('wbc count' in key.lower()):
					experiment.wbc_count= data[key][i]		
				elif ('# mononuclear cells' in key.lower()):
					experiment.mononuclear_num= data[key][i]
				elif ('# lymphocytes' in key.lower()):
					experiment.lymphocytes_num= data[key][i]
				elif ('% lymphocytes' in key.lower()):
					experiment.lymphocytes_per= data[key][i]
				elif ('# monocytes' in key.lower()):
					if('2' in key.lower()):
						experiment.monocytes2_num= data[key][i]
					else:
						experiment.monocytes_num= data[key][i]
				elif ('# neutrophils' in key.lower()):
					experiment.neutrophils_num= data[key][i]
				elif ('% neutrophils' in key.lower()):
					experiment.neutrophils_per= data[key][i]
				elif ('# eosinophils' in key.lower()):
					experiment.eosinophils_num= data[key][i]
				elif ('% eosinophils' in key.lower()):
					experiment.eosinophils_per= data[key][i]
				elif ('# basophils' in key.lower()):
					experiment.basophils_num= data[key][i]
				elif ('% basophils' in key.lower()):
					experiment.basophils_per= data[key][i]
				elif ('rbc count' in key.lower()):
					experiment.rbc_count= data[key][i]
				elif ('hematocrit' in key.lower()):
					experiment.ht= data[key][i]
				elif ('hemoblobin' in key.lower()):
					experiment.hb= data[key][i]
				elif ('plt platelet' in key.lower()):
					experiment.plt_count= data[key][i]
				elif ('platelet dist' in key.lower()):
					experiment.platelet_dist_range= data[key][i]
				elif ('platelet count' in key.lower()):
					experiment.platelet_count= data[key][i]
				elif ('average volume' in key.lower()):
					experiment.avg_vol_platelets= data[key][i]
				elif ('MCV' == key):
					experiment.mcv= data[key][i]
				elif ('RDV' == key):
					experiment.rdv= data[key][i]
				elif ('MCHC' == key):
					experiment.mchc= data[key][i]	
				elif ('MCV2' == key):
					experiment.mcv2= data[key][i]					
		if(flag):
			experiment.save()

def data_hem02(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Hem01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' == key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					experiment.timepoint= data[key][i]
				elif ('measurement date' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				elif ('sample id' in key.lower()):
					experiment.sample_id= data[key][i]
				elif ('wbc' in key.lower()):
					experiment.wbc_count= data[key][i]		
				elif ('mon #' in key.lower()):
					experiment.mononuclear_num= data[key][i]
				elif ('lym #' in key.lower()):
					experiment.lymphocytes_num= data[key][i]
				elif ('% lymphocytes' in key.lower()):
					experiment.lymphocytes_per= data[key][i]
				elif ('# monocytes' in key.lower()):
					if('2' in key.lower()):
						experiment.monocytes2_num= data[key][i]
					else:
						experiment.monocytes_num= data[key][i]
				elif ('# neutrophils' in key.lower()):
					experiment.neutrophils_num= data[key][i]
				elif ('% neutrophils' in key.lower()):
					experiment.neutrophils_per= data[key][i]
				elif ('# eosinophils' in key.lower()):
					experiment.eosinophils_num= data[key][i]
				elif ('% eosinophils' in key.lower()):
					experiment.eosinophils_per= data[key][i]
				elif ('# basophils' in key.lower()):
					experiment.basophils_num= data[key][i]
				elif ('% basophils' in key.lower()):
					experiment.basophils_per= data[key][i]
				elif ('rbc count' in key.lower()):
					experiment.rbc_count= data[key][i]
				elif ('hematocrit' in key.lower()):
					experiment.ht= data[key][i]
				elif ('hemoblobin' in key.lower()):
					experiment.hb= data[key][i]
				elif ('plt platelet' in key.lower()):
					experiment.plt_count= data[key][i]
				elif ('platelet dist' in key.lower()):
					experiment.platelet_dist_range= data[key][i]
				elif ('platelet count' in key.lower()):
					experiment.platelet_count= data[key][i]
				elif ('average volume' in key.lower()):
					experiment.avg_vol_platelets= data[key][i]
				elif ('MCV' == key):
					experiment.mcv= data[key][i]
				elif ('RDV' == key):
					experiment.rdv= data[key][i]
				elif ('MCHC' == key):
					experiment.mchc= data[key][i]	
				elif ('MCV2' == key):
					experiment.mcv2= data[key][i]					
		if(flag):
			experiment.save()

def data_hem01v2(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Hem01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' == key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					experiment.timepoint= data[key][i]
				elif ('measurement date' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				elif ('sample id' in key.lower()):
					experiment.sample_id= data[key][i]
				elif ('wbc' in key.lower()):
					experiment.wbc_count= data[key][i]		
				elif ('neu #' in key.lower()):
					experiment.neutrophils_num= data[key][i]
				elif ('lym #' in key.lower()):
					experiment.lymphocytes_num= data[key][i]
				elif ('mon #' in key.lower()):
					experiment.monocytes_num= data[key][i]
				elif ('eos #' in key.lower()):
					experiment.eosinophils_num= data[key][i]
				elif ('bas #' in key.lower()):
					experiment.basophils_num= data[key][i]
				elif ('neu %' in key.lower()):
					experiment.neutrophils_per= data[key][i]
				elif ('lym %' in key.lower()):
					experiment.lymphocytes_per= data[key][i]
				elif ('eos %' in key.lower()):
					experiment.eosinophils_per= data[key][i]
				elif ('bas %' in key.lower()):
					experiment.basophils_per= data[key][i]
				elif ('rbc' in key.lower()):
					experiment.rbc_count= data[key][i]
				elif ('HCT' in key):
					experiment.ht= data[key][i]
				elif ('HGB' in key):
					experiment.hb= data[key][i]
				elif ('MCV' in key):
					experiment.mcv= data[key][i]
				elif ('MCH (pg)' == key):
					experiment.mch= data[key][i]
				elif ('MCHC' in key):
					experiment.mchc= data[key][i]	
				elif ('RDW-CV' in key):
					experiment.rdwcv= data[key][i]
				elif ('RDW-SD' in key):
					experiment.rdwsd= data[key][i]
				elif ('PLT' in key):
					experiment.plt= data[key][i]
				elif ('MPV' in key):
					experiment.mpv= data[key][i]	
				elif ('PDW' in key):
					experiment.pdw= data[key][i]
				elif ('PCT' in key):
					experiment.pct= data[key][i]		
		if(flag):
			experiment.save()

def data_biochem01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Biochem01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sample source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('total protein' in key.lower()):
					experiment.total_protein = data[key][i]
				elif ('albumin' in key.lower()):
					experiment.albumin = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_biochem02(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Biochem02()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sample source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('sodium' in key.lower()):
					experiment.sodium = data[key][i]
				elif ('potassium' in key.lower()):
					experiment.potassium = data[key][i]
				elif ('chloride' in key.lower()):
					experiment.chloride = data[key][i]
				elif ('phosphorus' in key.lower()):
					experiment.phosphorus = data[key][i]
				elif ('calcium' in key.lower()):
					experiment.calcium = data[key][i]
				elif ('magnesium' in key.lower()):
					experiment.magnesium = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_biochem03(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Biochem03()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sample source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('ALT' in key):
					experiment.alt = data[key][i]
				elif ('ALP' in key):
					experiment.alp = data[key][i]
				elif ('AST' in key):
					experiment.ast = data[key][i]
				elif ('bilirubin total' in key.lower()):
					experiment.total_bilirubin = data[key][i]
				elif ('bilirubin direct' in key.lower()):
					experiment.direct_bilirubin = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_biochem04(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Biochem04()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sample source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('urea' in key.lower()):
					experiment.urea = data[key][i]
				elif ('uric acid' in key.lower()):
					experiment.uric_acid = data[key][i]
				elif ('creatinine' in key.lower()):
					experiment.creatinine = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_biochem05(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Biochem05()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sample source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('amylase' in key.lower()):
					experiment.amylase = data[key][i]
				elif ('lipase' in key.lower()):
					experiment.lipase = data[key][i]
				elif ('glucose' in key.lower()):
					experiment.glucose = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_biochem06(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Biochem06()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sample source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('cholesterol' in key.lower()):
					if('hdl' in key.lower()):
						experiment.hdl_cholesterol = data[key][i]
					elif('ldl' in key.lower()):
						experiment.ldl_cholesterol = data[key][i]
					else:
						experiment.cholesterol = data[key][i]
				elif ('triglycerides' in key.lower()):
					experiment.triglycerides = data[key][i]
				elif ('nefa' in key.lower()):
					experiment.nefa = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_biochem07(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Biochem07()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sample source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('iron' in key.lower()):
					experiment.iron = data[key][i]
				elif ('uibc' in key.lower()):
					experiment.uibc = data[key][i]
				elif ('ferritin' in key.lower()):
					experiment.ferritin = data[key][i]
				elif ('transferrin' in key.lower()):
					experiment.transferrin = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_biochem08(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Biochem08()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sample source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('ldl' in key.lower()):
					experiment.ldl = data[key][i]
				elif ('creatinine' in key.lower()):
					experiment.creatinine_kinase = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_hpibd02(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Hpibd02()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' == key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					experiment.timepoint= data[key][i]
				elif ('measurement date' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				elif ('inflammation score' in key.lower()):
					if('small' in key.lower()):
						experiment.inflammation_small= data[key][i]
					else:
						experiment.inflammation_large= data[key][i]
				elif ('epithelial damage' in key.lower()):
					if('small' in key.lower()):
						experiment.epithelial_small= data[key][i]
					else:
						experiment.epithelial_large= data[key][i]
				elif ('villus/crypt' in key.lower()):
					experiment.vc_ratio= data[key][i]
				elif ('colon length' in key.lower()):
					experiment.colon_length= data[key][i]
				elif ('epithelium flattening' in key.lower()):
					experiment.epithelium_s_flattening= data[key][i]
				elif ('apoptotic bodies' in key.lower()):
					if('small' in key.lower()):
						experiment.apoptotic_s_field= data[key][i]
					else:
						experiment.apoptotic_l_field= data[key][i]
				elif ('goblet cell' in key.lower()):
					if('small' in key.lower()):
						experiment.goblet_s_depletion= data[key][i]
					else:
						experiment.goblet_l_depletion= data[key][i]
				elif ('lumen exf' in key.lower()):
					if('small' in key.lower()):
						experiment.lumen_s_exfoliation= data[key][i]
					else:
						experiment.lumen_l_exfoliation= data[key][i]
				elif ('villus short' in key.lower()):
					experiment.villus_s_short= data[key][i]
				elif ('intestinal gland' in key.lower()):
					experiment.igd_small= data[key][i]
				elif ('paneth activation' in key.lower()):
					experiment.paneth_s_activation= data[key][i]
				elif ('peyer patch' in key.lower()):
					experiment.peyer_s_patch= data[key][i]
				elif ('crypt damage' in key.lower()):
					experiment.crypt_l_damage= data[key][i]
				elif ('endothelial' in key.lower()):
					experiment.endothelial_l_activation= data[key][i]
		if(flag):
			experiment.save()

######################################################
def data_hpni01(data,assay):
	print('inside hpni')
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Hpni01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('average' in key.lower()):
					if('gray' in key.lower()):
						experiment.avg_score_gray= data[key][i]
					elif('white'):
						experiment.avg_score_white= data[key][i]
				elif ('total score' in key.lower()):
					experiment.total_score = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				print(experiment)
		if(flag):				
			experiment.save()

def data_fc01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Fc01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('aquis' in key.lower()):
					experiment.live_aquis= data[key][i]					 
				elif ('total cell' in key.lower()):
					experiment.total_cell_count = data[key][i]
				elif ('% neu' in key.lower()):
					experiment.neu_per = data[key][i]					 
				elif ('% eoc' in key.lower()):
					experiment.eos_per = data[key][i]					 
				elif ('% mon' in key.lower()):
					experiment.mon_per = data[key][i]					 
				elif ('recent emigrant' in key.lower()):
					experiment.recent_emigrant_monocytes_per = data[key][i]
				elif ('inflammatory' in key.lower()):	
					experiment.inflammatory_monocytes_per = data[key][i]			
				elif ('steady state' in key.lower()):	
					experiment.steady_state_monocytes_per = data[key][i]

				elif ('nk cells' in key.lower()):	
					experiment.nk_cells_per = data[key][i]
				elif ('t cells' in key.lower()):	
					experiment.t_cells_per = data[key][i]
				elif ('b1b cells' in key.lower()):	
					experiment.b1b_cells_per = data[key][i]
				elif ('b2b cells' in key.lower()):	
					experiment.b2b_cells_per = data[key][i]

				elif ('% dcs' in key.lower()):
					experiment.dcs_per = data[key][i]
				elif ('% pdcs' in key.lower()):
					experiment.pdcs_per = data[key][i]
				elif ('% b cells' in key.lower()):
					experiment.b_cells_per = data[key][i]					 
				elif ('% cds' in key.lower()):
					if('type' in key.lower()):
						experiment.cds_cd11_per = data[key][i]	
					else:
						experiment.cds_per = data[key][i]					 
				elif ('macrophages' in key.lower()):
					experiment.macrophages_per = data[key][i]					 
									 
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				print(experiment)
		if(flag):				
			experiment.save()


def data_fc04(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Fc04()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('sample id' in key.lower()):
					experiment.sample_id = data[key][i]
				elif ('source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('aquis' in key.lower()):
					experiment.live_aquis= data[key][i]					 
				elif ('total cell' in key.lower()):
					experiment.total_cell_count = data[key][i]
				elif ('% all cells' in key.lower()):
					experiment.all_cells_per = data[key][i]
				elif ('% live' in key.lower()):
					if('leukocytes' in key.lower()):
						experiment.live_leukocytes_per = data[key][i]
					elif('epithelial'in key.lower()):	
						experiment.live_epithelial_per = data[key][i]
					else: experiment.live_cells_per = data[key][i]
				elif ('% early' in key.lower()):
					experiment.early_apoptotic_per = data[key][i]
				elif ('% late' in key.lower()):
					experiment.late_apoptotic_per = data[key][i]
				elif ('% neucrotic' in key.lower()):
					if('leukocytes' in key.lower()):
						experiment.total_leukocytes_per = data[key][i]
					elif('epithelial'in key.lower()):	
						experiment.total_epithelial_per = data[key][i]
					else: experiment.necrotic_per = data[key][i]
				elif ('# all cells' in key.lower()):
					experiment.all_cells_num = data[key][i]
				elif ('# live' in key.lower()):
					if('leukocytes' in key.lower()):
						experiment.live_leukocytes_num = data[key][i]
					elif('epithelial'in key.lower()):	
						experiment.live_epithelial_num = data[key][i]
					else: experiment.live_cells_num = data[key][i]
				elif ('# early' in key.lower()):
					experiment.early_apoptotic_num = data[key][i]
				elif ('# late' in key.lower()):
					experiment.late_apoptotic_num = data[key][i]
				elif ('# neucrotic' in key.lower()):
					if('leukocytes' in key.lower()):
						experiment.necrotic_leukocytes_num = data[key][i]
					elif('epithelial'in key.lower()):	
						experiment.necrotic_epithelial_num = data[key][i]
					else: experiment.necrotic_num = data[key][i]					 
				elif ('% total' in key.lower()):
					if('leukocytes' in key.lower()):
						experiment.total_leukocytes_per = data[key][i]
					elif('epithelial'in key.lower()):	
						experiment.total_epithelial_per = data[key][i]
				elif ('% apoptotic' in key.lower()):
					if('leukocytes' in key.lower()):
						experiment.total_leukocytes_per = data[key][i]
					elif('epithelial'in key.lower()):	
						experiment.total_epithelial_per = data[key][i]
				elif ('# total' in key.lower()):
					if('leukocytes' in key.lower()):
						experiment.total_leukocytes_num = data[key][i]
					elif('epithelial'in key.lower()):	
						experiment.total_epithelial_num = data[key][i]
				elif ('# apoptotic' in key.lower()):
					if('leukocytes' in key.lower()):
						experiment.total_leukocytes_num = data[key][i]
					elif('epithelial'in key.lower()):	
						experiment.total_epithelial_num = data[key][i]

				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				print(experiment)
		if(flag):				
			experiment.save()

def data_fc03(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Fc03()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('aquis' in key.lower()):
					experiment.live_aquis= data[key][i]					 
				elif ('total cell' in key.lower()):
					experiment.total_cell_count = data[key][i]
				elif ('% total B cells' in key.lower()):
					experiment.total_b_cells = data[key][i]
				elif ('b1b cells' in key.lower()):	
					experiment.b1b_cells_per = data[key][i]
				elif ('b2b cells' in key.lower()):	
					experiment.b2b_cells_per = data[key][i]
				elif ('immature' in key.lower()):	
					experiment.b2b_immature_cells_per = data[key][i]
				elif ('t1 cells' in key.lower()):	
					experiment.t1_cells_per = data[key][i]
				elif ('t2 cells' in key.lower()):	
					experiment.t2_cells_per = data[key][i]
				elif ('t3' in key.lower()):	
					experiment.t3_cells_per = data[key][i]
				elif ('mzb cells' in key.lower()):	
					experiment.mzb_cells_per = data[key][i]
				elif ('b2 mature' in key.lower()):	
					experiment.b2_mature_cells_per = data[key][i]		
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				print(experiment)
		if(flag):				
			experiment.save()

def data_fc07(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Fc07()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('facs' in key.lower()):
					experiment.facs_lysing= data[key][i]					 
				elif ('aquis' in key.lower()):
					experiment.live_aquis= data[key][i]					 
				elif ('total cell' in key.lower()):
					experiment.total_cell_count = data[key][i]
				elif ('% neu' in key.lower()):
					experiment.neu_per = data[key][i]					 
				elif ('% eos' in key.lower()):
					experiment.eos_per = data[key][i]					 
				elif ('% mon' in key.lower()):
					experiment.mon_per = data[key][i]					 
				elif ('% ly6chimono' in key.lower()):
					experiment.ly6chimono_per = data[key][i]					 
				elif ('% ly6cintermono' in key.lower()):
					experiment.ly6cintermono_per = data[key][i]					 
				elif ('% ly6clowmono' in key.lower()):
					experiment.ly6clowmono_per = data[key][i]					 
#				elif ('% dcs' in key.lower()):
#					experiment.dcs_per = data[key][i]
				elif ('% b cells' in key.lower()):
					experiment.b_cells_per = data[key][i]					 
				elif ('% cd8' in key.lower()):
					experiment.cd8_t_per = data[key][i]					 
				elif ('% cd4' in key.lower()):
					experiment.cd4_t_per = data[key][i]					 

				elif ('# neu' in key.lower()):
					experiment.neu_per = data[key][i]					 
				elif ('# eos' in key.lower()):
					experiment.eos_per = data[key][i]					 
				elif ('# mon' in key.lower()):
					ly6chimono_perexperiment.mon_per = data[key][i]					 
				elif ('# ly6chimono' in key.lower()):
					experiment.ly6chimono_per = data[key][i]					 
				elif ('# ly6cintermono' in key.lower()):
					experiment.ly6cintermono_per = data[key][i]					 
				elif ('# ly6clowmono' in key.lower()):
					experiment.ly6clowmono_per = data[key][i]
				elif ('# b cells' in key.lower()):
					experiment.b_cells_num = data[key][i]					 
				elif ('# cd8' in key.lower()):
					experiment.cd8_t_num = data[key][i]					 
				elif ('# cd4' in key.lower()):
					experiment.cd4_t_num = data[key][i]											 
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				print(experiment)
		if(flag):				
			experiment.save()

def data_fc08(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Fc08()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					#print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d
				elif ('sample id' in key.lower()):
					experiment.sample_id= data[key][i]
				elif ('source' in key.lower()):
					experiment.sample_source= data[key][i]
				elif ('facs' in key.lower()):
					experiment.facs_lysing= data[key][i]					 
				elif ('aquis' in key.lower()):
					experiment.live_aquis= data[key][i]					 
				elif ('total cell' in key.lower()):
					if('2' in key.lower()):
						experiment.total_cell_count_2 = data[key][i]
					else:
						experiment.total_cell_count = data[key][i]
				elif ('% neu' in key.lower()):
					experiment.neu_per = data[key][i]					 
				elif ('% eos' in key.lower()):
					experiment.eos_per = data[key][i]					 
				elif ('% mon' in key.lower()):
					experiment.mon_per = data[key][i]					 
				elif ('% ly6chimono' in key.lower()):
					experiment.ly6chimono_per = data[key][i]					 
				elif ('% ly6cintermono' in key.lower()):
					experiment.ly6cintermono_per = data[key][i]					 
				elif ('% ly6clowmono' in key.lower()):
					experiment.ly6clowmono_per = data[key][i]					 
				elif ('% dcs' in key.lower()):
					experiment.dcs_per = data[key][i]					 
				elif ('# neu' in key.lower()):
					experiment.neu_per = data[key][i]					 
				elif ('# eos' in key.lower()):
					experiment.eos_per = data[key][i]					 
				elif ('# mon' in key.lower()):
					ly6chimono_perexperiment.mon_per = data[key][i]					 
				elif ('# ly6chimono' in key.lower()):
					experiment.ly6chimono_per = data[key][i]					 
				elif ('# ly6cintermono' in key.lower()):
					experiment.ly6cintermono_per = data[key][i]					 
				elif ('# ly6clowmono' in key.lower()):
					experiment.ly6clowmono_per = data[key][i]					 
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
				print(experiment)
		if(flag):				
			experiment.save()
'''
    nk_cells_per = models.FloatField(blank=True, null=True)
    b_cells_per = models.FloatField(blank=True, null=True)
    t_cells_per = models.FloatField(blank=True, null=True)
    neu_num = models.FloatField(blank=True, null=True)
    eos_num = models.FloatField(blank=True, null=True)
    mon_num = models.FloatField(blank=True, null=True)
    ly6chimono_num = models.FloatField(blank=True, null=True)
    ly6cintermono_num = models.FloatField(blank=True, null=True)
    ly6clowmono_num = models.FloatField(blank=True, null=True)
    dcs_num = models.FloatField(blank=True, null=True)
    nk_cells_num = models.FloatField(blank=True, null=True)
    b_cells_num = models.FloatField(blank=True, null=True)
    t_cells_num = models.FloatField(blank=True, null=True)
    total_cell_count_2 = models.FloatField(blank=True, null=True)
    b_cells_per_2 = models.FloatField(blank=True, null=True)
    nk_cells_per_2 = models.FloatField(blank=True, null=True)
    cd4_per = models.FloatField(blank=True, null=True)
    cd4_cd25_per =models.FloatField(blank=True, null=True)


'''

def data_hpa02(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Hpa02()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('synovitis' in key.lower()):
					experiment.synovitis= data[key][i]
				elif ('cartilage destruction' in key.lower()):
					experiment.cartilage_destruction= data[key][i]
				elif ('bone erosion' in key.lower()):
					experiment.bone_erosion= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_ar02(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Ar02()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('parameter' in key.lower()):
					if ('1' in key.lower()):
						experiment.parameter1 = data[key][i]
					elif ('2' in key.lower()):
						experiment.parameter2 = data[key][i]
					elif ('3' in key.lower()):
						experiment.parameter3 = data[key][i]
					elif ('4' in key.lower()):
						experiment.parameter4 = data[key][i]					
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_ar03(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Ar03()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sum intensity' in key.lower()):
					if ('fl' in key.lower()):
						experiment.sum_intensity_fl = data[key][i]
					elif ('hl' in key.lower()):
						experiment.sum_intensity_hl = data[key][i]
					elif ('total' in key.lower()):
						experiment.sum_intensity_total = data[key][i]
				elif ('net intensity' in key.lower()):
					if ('fl' in key.lower()):
						experiment.net_intensity_fl = data[key][i]
					elif ('hl' in key.lower()):
						experiment.net_intensity_hl = data[key][i]
					elif ('total' in key.lower()):
						experiment.net_intensity_total = data[key][i]					
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_ar04(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Ar04()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('sum intensity' in key.lower()):
					experiment.sum_intensity= data[key][i]
				elif ('net intensity' in key.lower()):
					experiment.net_intensity = data[key][i]
				elif ('mean intensity' in key.lower()):
					experiment.mean_intensity = data[key][i]
				elif ('acquisition' in key.lower()):
					experiment.acquisition_settings = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_ar05(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Ar05()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
				elif ('BV/TV' in key):
					experiment.bvtv = data[key][i]
				elif ('BS/TV' in key):
					experiment.bstv= data[key][i]				
				elif ('trabecular' in key.lower()):
					if ('thickness' in key.lower()):
						experiment.trabecular_thickness= data[key][i]
					elif ('number' in key.lower()):
						experiment.trabecular_number= data[key][i]
					elif ('seperation' in key.lower()):
						experiment.trabecular_seperation= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_ar06(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Ar06()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('clinical' in key.lower()):
					experiment.clinical_score= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_ar07(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Ar07()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('weight' in key.lower()):
					experiment.weight= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_pr02(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Pr02()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('mg feacal mass' in key.lower()):
					experiment.mg_feacal_mass= data[key][i]
				elif ('concentration' in key.lower()):
					experiment.protein_concentration= data[key][i]
				elif ('protocol' in key.lower()):
					experiment.digestion_protocol= data[key][i]
				elif ('amount' in key.lower()):
					experiment.amount_injected= data[key][i]
				elif ('LC-MS/MS' in key):
					experiment.lc_ms_analysis_time= data[key][i]
				elif ('instrument' in key.lower()):
					experiment.instrument_method= data[key][i]
				elif ('total l bacterial' in key.lower()):
					experiment.total_bac_signal= data[key][i]
				elif ('firmicutes' in key.lower()):
					if ('signal' in key.lower()):
						experiment.firmicutes_per = data[key][i]
					elif('ratio' in key.lower()):
						experiment.firm_bac_ratio = data[key][i]					
				elif ('bacteroidetes' in key.lower()):
					experiment.bacteroidetes_per= data[key][i]
				elif ('proteobacteria' in key.lower()):
					experiment.proteobacteria_per= data[key][i]					
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_cba01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Cba01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('dilution factor' in key.lower()):
					experiment.dilution_factor = data[key][i]
				elif ('il-6' in key.lower()):
					experiment.il_six = data[key][i]
				elif ('il-10' in key.lower()):
					experiment.il_ten = data[key][i]
				elif ('il-12' in key.lower()):
					experiment.il_twelve = data[key][i]
				elif ('il-17' in key.lower()):
					experiment.il_seventeen = data[key][i]
				elif ('ifn' in key.lower()):
					experiment.ifn_gamma = data[key][i]
				elif ('tnf' in key.lower()):
					experiment.tnf_alpha = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_cba02(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Cba02()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('dilution factor' in key.lower()):
					experiment.dilution_factor = data[key][i]
				elif ('igg1 ng/ml' in key.lower()):
					experiment.igg1 = data[key][i]
				elif ('igg2a ng/ml' in key.lower()):
					experiment.igg2a = data[key][i]
				elif ('igg3 ng/ml' in key.lower()):
					experiment.igg3 = data[key][i]
				elif ('iga' in key.lower()):
					experiment.iga = data[key][i]
				elif ('igm' in key.lower()):
					experiment.igm = data[key][i]
				elif ('igg2b' in key.lower()):
					experiment.igg2b = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_cba03(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Cba03()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('parameter' in key.lower()):
					if ('1' in key.lower()):
						experiment.parameter1 = data[key][i]
					elif ('2' in key.lower()):
						experiment.parameter2 = data[key][i]
					elif ('3' in key.lower()):
						experiment.parameter3 = data[key][i]
					elif ('4' in key.lower()):
						experiment.parameter4 = data[key][i]					
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_hpibd01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Hpibd01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('histology' in key.lower()):
					experiment.histology_num = data[key][i]
				elif ('colon' in key.lower()):
					experiment.colon_length = data[key][i]
				elif ('inflammation' in key.lower()):
					experiment.inflammation_score = data[key][i]
				elif ('tissue damage score' in key.lower()):
					experiment.tissue_damage_score = data[key][i]
				elif ('ulceration score' in key.lower()):
					experiment.ulceration_score = data[key][i]										
				elif ('involvement' in key.lower()):
					experiment.involvement_per = data[key][i]	
				elif ('total colitis score' in key.lower()):
					experiment.total_colitis_score = data[key][i]						
				elif ('tumor' in key.lower()):
					if('macroscopic' in key.lower()):
						experiment.tumor_num = data[key][i]
					elif ('size' in key.lower()):
						experiment.tumor_size = data[key][i]									
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_hpibd03(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Hpibd03()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('histology' in key.lower()):
					experiment.histology_num = data[key][i]
				elif ('si inflammation' in key.lower()):
					experiment.si_inflammation_score = data[key][i]
				elif ('colon inflammation' in key.lower()):
					experiment.colon_inflammation_score = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_hpibd04(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Hpibd04()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('slide' in key.lower()):
					experiment.slide_number = data[key][i]
				elif ('leukocyte' in key.lower()):
					if ('1' in key.lower()):
						experiment.leukocyte_density_1 = data[key][i]
					elif ('2' in key.lower()):
						experiment.leukocyte_density_2 = data[key][i]
					elif ('3' in key.lower()):
						experiment.leukocyte_density_3 = data[key][i]
					elif ('avg' in key.lower()):
						experiment.leukocyte_density_avg = data[key][i]
				elif ('level' in key.lower()):
					if ('1' in key.lower()):
						experiment.level_1 = data[key][i]
					elif ('2' in key.lower()):
						experiment.level_2 = data[key][i]
					elif ('3' in key.lower()):
						experiment.level_3 = data[key][i]
					elif ('avg' in key.lower()):
						experiment.level_avg = data[key][i]
				elif ('extent' in key.lower()):
					if ('1' in key.lower()):
						experiment.extent_1 = data[key][i]
					elif ('2' in key.lower()):
						experiment.extent_2 = data[key][i]
					elif ('3' in key.lower()):
						experiment.extent_3 = data[key][i]
					elif ('avg' in key.lower()):
						experiment.extent_avg = data[key][i]						
				elif ('inflammation' in key.lower()):
					experiment.inflammation_score = data[key][i]
				elif ('abscesses' in key.lower()):
					experiment.abscesses = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_endo01(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = Endo01()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					#print(data[key][i])
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					print(data[key][i])
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('thickening' in key.lower()):
					experiment.thicken_colon = data[key][i]
				elif ('vascular' in key.lower()):
					experiment.changes_vascular = data[key][i]
				elif ('fibrin' in key.lower()):
					experiment.fibrin_visible = data[key][i]
				elif ('granularity' in key.lower()):
					experiment.granularity_mucosak = data[key][i]
				elif ('stool consistency' in key.lower()):
					experiment.stool_consistency = data[key][i]
				elif ('total colitis score' in key.lower()):
					experiment.total_colitis_score = data[key][i]																		
				elif ('number of tumors' in key.lower()):
					if ('1' in key.lower()):
						experiment.num_tumor_size_1 = data[key][i]
					elif ('2' in key.lower()):
						experiment.num_tumor_size_2 = data[key][i]
					elif ('3' in key.lower()):
						experiment.num_tumor_size_3 = data[key][i]
					elif ('4' in key.lower()):
						experiment.num_tumor_size_4 = data[key][i]
					elif ('5' in key.lower()):
						experiment.num_tumor_size_5 = data[key][i]												
				elif ('total tumor' in key.lower()):
					experiment.total_tumor_num = data[key][i]
				elif ('tumor load' in key.lower()):
					experiment.tumor_load = data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_gi05(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = GI05()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('spleen weight' in key.lower()):
					experiment.spleen_weight = data[key][i]
				elif ('cfus/g spleen' in key.lower()):
					experiment.spleen_cfus= data[key][i]
				elif ('sample weight' in key.lower()):
					experiment.weight= data[key][i]
				elif ('fecal cfus' in key.lower()):
					experiment.fecal_cfus= data[key][i]
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

def data_gi06(data,assay):
	for i in range(len(data["Mouse ID"])):
		flag =1
		experiment = GI06()
		experiment.assayid = assay
		for key in data:
			if(flag):
				if ('Mouse ID' in key):
					#Return mouse id from mouse table
					if(data[key][i] is None):
						flag = 0
						break; 
					mouse = Mouse.objects.get(mid=data[key][i])
					if(mouse): 
						experiment.mid = mouse
					else:
						flag = 0
						break;			
				elif ('timepoint' in key.lower()):
					experiment.timepoint= data[key][i]
				elif ('Age' in key):
					experiment.age= data[key][i]
				elif ('measurement' in key.lower()):
					d = data[key][i]
					experiment.measurement_date= d				 
				elif ('firmicutes' in key.lower()):
					experiment.firmicutes= data[key][i]
				elif ('bacteroidetes' in key.lower()):
					experiment.bacteroidetes= data[key][i]
				elif ('proteobacteria' in key.lower()):
					experiment.proteobacteria= data[key][i]
				elif ('ratio' in key.lower()):
					experiment.firmicutes_bacteroidetes_ratio= data[key][i]										
				elif ('comment' in key.lower()):
					experiment.comment= data[key][i]
		if(flag):				
			experiment.save()

#Returns number of rows & number of columns with data in excel tab  
def find_edges(sheet):
    row = sheet.max_row
    while row > 0:
        cells = sheet[row]
        if all([cell.value is None for cell in cells]):
            row -= 1
        else:
            break
    if row == 0:
        return 0, 0

    column = sheet.max_column
    while column > 0:
        cells = next(sheet.iter_cols(min_col=column, max_col=column, max_row=row))
        if all([cell.value is None for cell in cells]):
            column -= 1
        else:
            break
    return row, column

def dataofAssay(assay,filename):
	wb = openpyxl.load_workbook(filename, data_only=True)
	print("OK")
	#Optional condition
	if(assay not in tableNames):
		print(assay)
		return -1;
	#For info tab to find the mouselist
	if(assay in "info"):
		sheet = wb[wb.sheetnames[0]]
	#For regular assay
	else:
		sheet = wb[assay]
	[rows,cols]=find_edges(sheet)
	#Column_C=sheet['Mouse ID']   
	#print ( len(column_C))
	#print("test")
	#rows = len(sheet["Mouse ID"])
	d = {}
	d = collections.OrderedDict()

	# If mouseID // mouse strain
	for c in range(1,cols):
		r0 = sheet.cell(row=1, column = c).value
		if(r0 and r0 not in d):
			d[r0] = []
		for r in range(2,rows+1):
			object = sheet.cell(row=r, column = c).value
			d[r0].append(object)
	return d;

def handle_uploaded_file(assayobject):
    #with open('some/file/name.txt', 'wb+') as destination:
    #    for chunk in f.chunks():
    #        destination.write(chunk)
	fname = '../static'+assayobject.rawdata_file.url
	wb = openpyxl.load_workbook(fname)
	info = dataofAssay("info",fname)
	#print(info)
	data = dataofAssay(assayobject.type.code,fname)
	if(data==-1):
		return -1
	print(data.keys())
	print("Data")
	print(data)
	#print(len(data["Mouse ID"]))
	create_mouseHash(info)
	print('\n\n\n\n\n')
	if(assayobject.type.code == "IINFLC-04"):
		data_iinflc04(data,assayobject)
	if(assayobject.type.code == "NI-01"):
		data_ni01(data,assayobject)
	if(assayobject.type.code == "NI-02-ROT-01"):
		data_ni02rot01(data,assayobject)
	if(assayobject.type.code == "NI-02-OFD-01"):
		data_ni02ofd01(data,assayobject)
	if(assayobject.type.code == "NI-02-GRS-01"):
		data_ni02grs01(data,assayobject)
	if(assayobject.type.code == "HEM-01"):
		data_hem01v2(data,assayobject)
	if(assayobject.type.code == "IINFLC-02"):
		data_iinflc02(data,assayobject)
	if(assayobject.type.code == "IINFLC-03"):
		data_iinflc03(data,assayobject)
	if(assayobject.type.code == "HPIBD-02"):
		data_hpibd02(data,assayobject)
	if(assayobject.type.code == "BIOCHEM-01"):
		data_biochem01(data,assayobject)
	if(assayobject.type.code == "BIOCHEM-02"):
		data_biochem02(data,assayobject)
	if(assayobject.type.code == "BIOCHEM-03"):
		data_biochem03(data,assayobject)
	if(assayobject.type.code == "BIOCHEM-04"):
		data_biochem04(data,assayobject)
	if(assayobject.type.code == "BIOCHEM-05"):
		data_biochem05(data,assayobject)
	if(assayobject.type.code == "BIOCHEM-06"):
		data_biochem06(data,assayobject)
	if(assayobject.type.code == "BIOCHEM-07"):
		data_biochem07(data,assayobject)
	if(assayobject.type.code == "BIOCHEM-08"):
		data_biochem08(data,assayobject)
	if(assayobject.type.code == "HPNI-01"):
		data_hpni01(data,assayobject)
	if(assayobject.type.code == "FC-07"):
		data_fc07(data,assayobject)
	if(assayobject.type.code == "FC-08"):
		data_fc08(data,assayobject)
	if(assayobject.type.code == "AR-02"):
		data_ar02(data,assayobject)
	if(assayobject.type.code == "AR-03"):
		data_ar03(data,assayobject)
	if(assayobject.type.code == "AR-04"):
		data_ar04(data,assayobject)
	if(assayobject.type.code == "AR-05"):
		data_ar05(data,assayobject)
	if(assayobject.type.code == "AR-06"):
		data_ar06(data,assayobject)
	if(assayobject.type.code == "AR-07"):
		data_ar07(data,assayobject)
	if(assayobject.type.code == "IINFLC-05"):
		data_iinflc05(data,assayobject)
	if(assayobject.type.code == "IINFLC-06"):
		data_iinflc06(data,assayobject)				
	if(assayobject.type.code == "PR-02"):
		data_pr02(data,assayobject)
	if(assayobject.type.code == "CBA-01"):
		data_cba01(data,assayobject)
	if(assayobject.type.code == "CBA-02"):
		data_cba02(data,assayobject)
	if(assayobject.type.code == "CBA-03"):
		data_cba03(data,assayobject)
	if(assayobject.type.code == "HPIBD-03"):
		data_hpibd03(data,assayobject)
	if(assayobject.type.code == "HPIBD-01"):
		data_hpibd01(data,assayobject)
	if(assayobject.type.code == "HPIBD-04"):
		data_hpibd04(data,assayobject)		
	if(assayobject.type.code == "ENDO-01"):
		data_endo01(data,assayobject)
	if(assayobject.type.code == "IINFLC-01"):
		data_iinflc01(data,assayobject)
	if(assayobject.type.code == "FC-01"):
		data_fc01(data,assayobject)
	if(assayobject.type.code == "FC-03"):
		data_fc03(data,assayobject)
	if(assayobject.type.code == "FC-04"):
		data_fc04(data,assayobject)
	if(assayobject.type.code == "HPA-02"):
		data_hpa02(data,assayobject)
	if(assayobject.type.code == "GI-05"):
		data_gi05(data,assayobject)
	if(assayobject.type.code == "GI-06"):
		data_gi06(data,assayobject)

def returnTemplateName(assayobject):
	switcher ={
		4:'assays/assaytypes/iinflc-03.html',
		5:'assays/assaytypes/iinflc-04.html',
		6:'assays/assaytypes/iinflc-02.html',
		7:'assays/assaytypes/ni01.html',
		8:'assays/assaytypes/ni02rot01.html',
		9:'assays/assaytypes/ni02ofd01.html',
		10:'assays/assaytypes/ni02grs01.html',
		11:'assays/assaytypes/hem01.html',
		12:'assays/assaytypes/hpibd-02.html',
		13:'assays/assaytypes/biochem01.html',
		14:'assays/assaytypes/biochem02.html',		
		15:'assays/assaytypes/biochem03.html',		
		16:'assays/assaytypes/biochem04.html',
		17:'assays/assaytypes/biochem05.html',		
		18:'assays/assaytypes/biochem06.html',
		19:'assays/assaytypes/biochem07.html',
		20:'assays/assaytypes/biochem08.html',
		22:'assays/assaytypes/hpni01.html',		
		23:'assays/assaytypes/fc08.html',
		24:'assays/assaytypes/ar02.html',
		25:'assays/assaytypes/iinflc-05.html',
		26:'assays/assaytypes/iinflc-06.html',
		27:'assays/assaytypes/fc07.html',
		28:'assays/assaytypes/pr02.html',
		29:'assays/assaytypes/cba01.html',	
		30:'assays/assaytypes/cba02.html',
		31:'assays/assaytypes/hpibd03.html',
		32:'assays/assaytypes/hpibd01.html',
		33:'assays/assaytypes/hpibd04.html',
		34:'assays/assaytypes/endo01.html',
		35:'assays/assaytypes/iinflc-01.html',
		36:'assays/assaytypes/ar03.html',
		37:'assays/assaytypes/ar04.html',
		38:'assays/assaytypes/ar05.html',
		39:'assays/assaytypes/ar06.html',
		40:'assays/assaytypes/ar07.html',
		41:'assays/assaytypes/cba03.html',
		42:'assays/assaytypes/fc01.html',
		43:'assays/assaytypes/fc03.html',
		44:'assays/assaytypes/hpa02.html',
		45:'assays/assaytypes/fc04.html',
		46:'assays/assaytypes/gi05.html',
		47:'assays/assaytypes/gi06.html',
	}
	return switcher.get(assayobject.id,"Ivalid")

def get_parameters(assay):
	switcher ={
		4: Iinflc03._meta.get_fields(),
		5: Iinflc04._meta.get_fields(),
		6: Iinflc02._meta.get_fields(),
		7: Ni01._meta.get_fields(),
		8: Ni02Rot01._meta.get_fields(),
		9: Ni02ofd01._meta.get_fields(),
		10: Ni02grs01._meta.get_fields(),
		11: Hem01._meta.get_fields(),
		12: Hpibd02._meta.get_fields(),
		13: Biochem01._meta.get_fields(),
		14: Biochem02._meta.get_fields(),		
		15: Biochem03._meta.get_fields(),		
		16: Biochem04._meta.get_fields(),
		17: Biochem05._meta.get_fields(),		
		18: Biochem06._meta.get_fields(),
		19: Biochem07._meta.get_fields(),
		20: Biochem08._meta.get_fields(),
		22: Hpni01._meta.get_fields(),		
		23: Fc08._meta.get_fields(),
		24: Ar02._meta.get_fields(),
		25: Iinflc05._meta.get_fields(),
		26: Iinflc06._meta.get_fields(),
		27: Fc07._meta.get_fields(),
		28: Pr02._meta.get_fields(),
		29: Cba01._meta.get_fields(),
		30: Cba02._meta.get_fields(),
		31: Hpibd03._meta.get_fields(),
		32: Hpibd01._meta.get_fields(),
		33: Hpibd04._meta.get_fields(),
		34: Endo01._meta.get_fields(),
		35: Iinflc01._meta.get_fields(),
		36: Ar03._meta.get_fields(),
		37: Ar04._meta.get_fields(),
		38: Ar05._meta.get_fields(),
		39: Ar06._meta.get_fields(),
		40: Ar07._meta.get_fields(),
		41: Cba03._meta.get_fields(),
		42: Fc01._meta.get_fields(),
		43: Fc03._meta.get_fields(),
		44: Hpa02._meta.get_fields(),
		45: Fc04._meta.get_fields(),
		46: GI05._meta.get_fields(),
		47: GI06._meta.get_fields()
	}
	parameters = []
	parameters_names = []
	pars = switcher.get(assay.type.id,"Ivalid")
	if (pars):
		for i in range(6,len(pars)-1):
			if(pars[i].get_internal_type()=='FloatField' or pars[i].get_internal_type()=='IntegerField'):
				print("check----------------------------------------------------------------------------")
				parameters.append(pars[i].name)
				parameters_names.append(pars[i].verbose_name)
	return parameters, parameters_names

def parameterMeasures(measures, parameter):
#    print('Inside Parameter Measures')
    flag = True
    mouselist = measures.values('mid').distinct().order_by('mid')
    gender = Mouse.objects.filter(id__in=mouselist).values_list('gender', flat=True).exclude(gender__isnull=True).distinct()
    genotype = Mouse.objects.filter(id__in=mouselist).values_list('genotype', flat=True).exclude(genotype__isnull=True).distinct()
    induced = Mouse.objects.filter(id__in=mouselist).values_list('induced', flat=True).exclude(induced__isnull=True).distinct()
    treated = Mouse.objects.filter(id__in=mouselist).values_list('treated', flat=True).exclude(treated__isnull=True).distinct()
#    print('gender {} {}',genotype, len(genotype))
#    print('sex {} {}',gender, len(genotype))
#    print('induced {} {}',induced, len(induced))
#    print('treated {} {}',treated, len(treated))
    isnull=False
    test = {}
    if(gender.exists()):
#        print('gender not empty')
        if(genotype.exists()):
#            print('genotype not empty')
            for sex in gender:
            	if(sex):
                    for gene in genotype:
                        if(gene):
                            label = sex + " "+gene
                            m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex)
                            parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                            df222 = pd.DataFrame(list(parameter_measures.values(parameter)))
#                            print("Parameter Measures")
#                            print(all(df222))
                            flag = False
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
#                            print('Test dataframe parameter')
                            cols_to_check = df.columns
                            df['is_na'] = df[cols_to_check].isnull().apply(lambda x: all(x), axis=1)
                            if(df['is_na'].all()==True):
                            	isnull=True
                            	break;
                            '''if df['timepoint'].empty:
                                print('Empty dataframe')
                            else:
                            	print('Not empty dataframe')
                            	df['timepoint_is'] = map(lambda x: x.isdigit(), df['timepoint'])
                            	print(df['timepoint_is'].empty)
                            	'''
                            #par_df = df[parameter]
                            #if par_df.empty:
                            #    print('par_df')
                            if not df.empty:
                                df = df[['timepoint',parameter]] 
#                            print(df)
                            test[label] = df
                        else:
                            label = sex
                            m = Mouse.objects.filter(id__in= mouselist,gender=sex)
                            parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                            if not df.empty:
                                df = df[['timepoint',parameter]] 
#                            print(df)
                            test[label] = df
    if(isnull):
    	test = {}
#    print('Out Parameter Measures')
    return test,flag,genotype

def parameterMeasures2(measures, parameter,all):
    print("Version 2")
    flag = False
    mouselist = measures.values('mid').distinct().order_by('mid')
    gender = Mouse.objects.filter(id__in=mouselist).values_list('gender', flat=True).exclude(gender__isnull=True).distinct()
    genotype = Mouse.objects.filter(id__in=mouselist).values_list('genotype', flat=True).exclude(genotype__isnull=True).distinct()
    induced = Mouse.objects.filter(id__in=mouselist).values_list('induced', flat=True).exclude(induced__isnull=True).distinct()
    treated = Mouse.objects.filter(id__in=mouselist).values_list('treated', flat=True).exclude(treated__isnull=True).distinct()
#    print('gender {} {}',genotype, len(genotype))
#    print('sex {} {}',gender, len(genotype))
#    print('induced {} {}',induced, len(induced))
#    print('treated {} {}',treated, len(treated))

    gendercount = findlabels(mouselist,1,gender,measures)
    genotypecount = findlabels(mouselist,2,genotype,measures)
    inducedcount = findlabels(mouselist,3,induced,measures)
    treatedcount = findlabels(mouselist,4,treated,measures)

#    print('gendercount {}', genotypecount)
#    print('sexcount {} ',gendercount)
#    print('inducedcount {}',inducedcount)
#    print('treatedcount {}',treatedcount)
    test = {}
    if(len(gender)>1 and gendercount == 1):
        for sex in gender:
            if(len(genotype)>1 and genotypecount == 1):
 #               print("First flag")
                for gene in genotype:
                    if(len(induced)>1 and inducedcount == 1):
                        for inducedlabel in induced:
                            if(len(treated)>1 and treatedcount == 1):
                                for treatedlabel in treated:
                                    #To DO
                                    m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex,induced= inducedlabel,treated=treatedlabel)
                                    parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                                    label = sex + " "+gene +" "+inducedlabel + " " + treatedlabel
                                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                                    df.dropna()
                                    test[label] = df
                            else:
                                m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex,induced= inducedlabel)
                                parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                                label = sex + " "+gene +" "+inducedlabel
                                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                                df.dropna()
                                test[label] = df
                    elif(len(treated)>1 and treatedcount == 1):
                        m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex,treated=treatedlabel)
                        parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                        label = sex + " "+gene +" "+treatedlabel
                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                        df.dropna()
                        test[label] = df                  	
                    else:
                        m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex)
                        parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                        label = sex + " "+gene
                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                        df.dropna()
                        test[label] = df
            elif(len(induced)>1 and inducedcount == 1):
                for inducedlabel in induced:
                    if(len(treated)>1 and treatedcount == 1):
                       for treatedlabel in treated:
                            m = Mouse.objects.filter(id__in= mouselist,gender=sex,induced= inducedlabel,treated=treatedlabel)
                            parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                            label = sex +" "+inducedlabel + " " + treatedlabel
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                            df.dropna()
                            test[label] = df
                    else:
                        m = Mouse.objects.filter(id__in= mouselist,gender=sex,induced= inducedlabel)
                        parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                        label = sex + " "+inducedlabel
                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                        df.dropna()
                        test[label] = df
            elif(len(treated)>1 and treatedcount == 1):
                m = Mouse.objects.filter(id__in= mouselist,gender=sex,treated=treatedlabel)
                parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                label = sex+" "+treatedlabel
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()  
                test[label] = df

            else:
                m = Mouse.objects.filter(id__in= mouselist,gender=sex)
                parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                label = sex
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
    elif(len(genotype)>1 and genotypecount == 1):
        for gene in genotype:
            if(len(induced)>1 and inducedcount == 1):
                for inducedlabel in induced:
                    if(len(treated)>1 and treatedcount == 1):
                        for treatedlabel in treated:
                            m = Mouse.objects.filter(id__in= mouselist,genotype=gene,induced= inducedlabel,treated=treatedlabel)
                            parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                            label = gene +" "+inducedlabel + " " + treatedlabel
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                            df.dropna()
                            test[label] = df
                    else:
                        m = Mouse.objects.filter(id__in= mouselist,genotype=gene,induced= inducedlabel)
                        parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                        label = gene +" "+inducedlabel
                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                        df.dropna()
                        test[label] = df
            elif(len(treated)>1 and treatedcount == 1):
                m = Mouse.objects.filter(id__in= mouselist,genotype=gene,treated=treatedlabel)
                parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                label = gene +" "+treatedlabel
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
            else:
                m = Mouse.objects.filter(id__in= mouselist,genotype=gene)
                parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                label = gene
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
    elif(len(induced)>1 and inducedcount == 1):
        for inducedlabel in induced:
            if(len(treated)>1 and treatedcount == 1):
                for treatedlabel in treated:
                    m = Mouse.objects.filter(id__in= mouselist,induced= inducedlabel,treated=treatedlabel)
                    parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                    label = inducedlabel + " " + treatedlabel
                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                    df.dropna()
                    test[label] = df
            else:
                m = Mouse.objects.filter(id__in= mouselist,induced= inducedlabel)
                parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
                label = inducedlabel
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
    elif(len(treated)>1 and treatedcount == 1):
    	for treatedlabel in treated:
            m = Mouse.objects.filter(id__in= mouselist,genotype=gene,treated=treatedlabel)
            parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
            label = treatedlabel
            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
            df.dropna()
            test[label] = df
    else:
        if(all==False):
            m = Mouse.objects.filter(id__in= mouselist)
            parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
            label = 'All'
            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
            df.dropna()
            test[label] = df
    if(all==True):
        m = Mouse.objects.filter(id__in= mouselist)
        parameter_measures = measures.filter(mid__in = m).values_list('timepoint',parameter)
        label = 'All'
        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
        df.dropna()
        test[label] = df   	
    if(df.empty):
    	flag=True
#    print('Out Parameter Measures')
    #clean_dict = filter(lambda k: not isnan(test[k]), test)
    return test,flag,genotype

def parameterMeasures4(measures, parameter,time_var):
    print("Version 4")
    flag = False
    mouselist = measures.values('mid').distinct().order_by('mid')
    gender = Mouse.objects.filter(id__in=mouselist).values_list('gender', flat=True).exclude(gender__isnull=True).distinct()
    genotype = Mouse.objects.filter(id__in=mouselist).values_list('genotype', flat=True).exclude(genotype__isnull=True).distinct()
    induced = Mouse.objects.filter(id__in=mouselist).values_list('induced', flat=True).exclude(induced__isnull=True).distinct()
    treated = Mouse.objects.filter(id__in=mouselist).values_list('treated', flat=True).exclude(treated__isnull=True).distinct()

    gendercount = findlabels(mouselist,1,gender,measures)
    genotypecount = findlabels(mouselist,2,genotype,measures)
    inducedcount = findlabels(mouselist,3,induced,measures)
    treatedcount = findlabels(mouselist,4,treated,measures)

    #timepoints = measures.values('timepoint').distinct().order_by('timepoint')
    
    test = {}
    if(len(gender)>1 and gendercount == 1):
        for sex in gender:
            if(len(genotype)>1 and genotypecount == 1):
                #print("First flag")
                for gene in genotype:
                    if(len(induced)>1 and inducedcount == 1):
                        for inducedlabel in induced:
                            if(len(treated)>1 and treatedcount == 1):
                                for treatedlabel in treated:
                                    #To DO
                                    m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex,induced= inducedlabel,treated=treatedlabel)
                                    parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                                    label = sex + " "+gene +" "+inducedlabel + " " + treatedlabel
                                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                                    df.dropna()
                                    test[label] = df
                            else:
                                m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex,induced= inducedlabel)
                                parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                                label = sex + " "+gene +" "+inducedlabel
                                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                                df.dropna()
                                test[label] = df
                    elif(len(treated)>1 and treatedcount == 1):
                        m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex,treated=treatedlabel)
                        parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                        label = sex + " "+gene +" "+treatedlabel
                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                        df.dropna()
                        test[label] = df                  	
                    else:
                        m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex)
                        parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                        label = sex + " "+gene
                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                        df.dropna()
                        test[label] = df
            elif(len(induced)>1 and inducedcount == 1):
                for inducedlabel in induced:
                    if(len(treated)>1 and treatedcount == 1):
                       for treatedlabel in treated:
                            m = Mouse.objects.filter(id__in= mouselist,gender=sex,induced= inducedlabel,treated=treatedlabel)
                            parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                            label = sex +" "+inducedlabel + " " + treatedlabel
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                            df.dropna()
                            test[label] = df
                    else:
                        m = Mouse.objects.filter(id__in= mouselist,gender=sex,induced= inducedlabel)
                        parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                        label = sex + " "+inducedlabel
                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                        df.dropna()
                        test[label] = df
            elif(len(treated)>1 and treatedcount == 1):
                m = Mouse.objects.filter(id__in= mouselist,gender=sex,treated=treatedlabel)
                parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                label = sex+" "+treatedlabel
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()  
                test[label] = df

            else:
                m = Mouse.objects.filter(id__in= mouselist,gender=sex)
                parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                label = sex
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
    elif(len(genotype)>1 and genotypecount == 1):
        for gene in genotype:
            if(len(induced)>1 and inducedcount == 1):
                for inducedlabel in induced:
                    if(len(treated)>1 and treatedcount == 1):
                        for treatedlabel in treated:
                            m = Mouse.objects.filter(id__in= mouselist,genotype=gene,induced= inducedlabel,treated=treatedlabel)
                            parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                            label = gene +" "+inducedlabel + " " + treatedlabel
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                            df.dropna()
                            test[label] = df
                    else:
                        m = Mouse.objects.filter(id__in= mouselist,genotype=gene,induced= inducedlabel)
                        parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                        label = gene +" "+inducedlabel
                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                        df.dropna()
                        test[label] = df
            elif(len(treated)>1 and treatedcount == 1):
                m = Mouse.objects.filter(id__in= mouselist,genotype=gene,treated=treatedlabel)
                parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                label = gene +" "+treatedlabel
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
            else:
                m = Mouse.objects.filter(id__in= mouselist,genotype=gene)
                parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                label = gene
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
    elif(len(induced)>1 and inducedcount == 1):
        for inducedlabel in induced:
            if(len(treated)>1 and treatedcount == 1):
                for treatedlabel in treated:
                    m = Mouse.objects.filter(id__in= mouselist,induced= inducedlabel,treated=treatedlabel)
                    parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                    label = inducedlabel + " " + treatedlabel
                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                    df.dropna()
                    test[label] = df
            else:
                m = Mouse.objects.filter(id__in= mouselist,induced= inducedlabel)
                parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
                label = inducedlabel
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
    elif(len(treated)>1 and treatedcount == 1):
        for treatedlabel in treated:
            m = Mouse.objects.filter(id__in= mouselist,genotype=gene,treated=treatedlabel)
            parameter_measures = measures.filter(mid__in = m,timepoint=time_var).values_list('timepoint',parameter)
            label = treatedlabel
            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
            df.dropna()
            test[label] = df
    if(df.empty):
    	flag=True
    print("Finish Version 4")
    return test,flag,genotype

def parameterMeasures3(measures, parameter,all):
    print("Version 3")
    flag = False
    mouselist = measures.values('mid').distinct().order_by('mid')
    gender = Mouse.objects.filter(id__in=mouselist).values_list('gender', flat=True).exclude(gender__isnull=True).distinct()
    genotype = Mouse.objects.filter(id__in=mouselist).values_list('genotype', flat=True).exclude(genotype__isnull=True).distinct()
    induced = Mouse.objects.filter(id__in=mouselist).values_list('induced', flat=True).exclude(induced__isnull=True).distinct()
    treated = Mouse.objects.filter(id__in=mouselist).values_list('treated', flat=True).exclude(treated__isnull=True).distinct()
    #print('gender {} {}',genotype, len(genotype))
    #print('sex {} {}',gender, len(genotype))
    #print('induced {} {}',induced, len(induced))
    #print('treated {} {}',treated, len(treated))

    gendercount = findlabels(mouselist,1,gender,measures)
    genotypecount = findlabels(mouselist,2,genotype,measures)
    inducedcount = findlabels(mouselist,3,induced,measures)
    treatedcount = findlabels(mouselist,4,treated,measures)

    timepoints = measures.values('timepoint').distinct().order_by('timepoint')
    #print('gendercount {}', genotypecount)
    #print('sexcount {} ',gendercount)
    #print('inducedcount {}',inducedcount)
    #print('treatedcount {}',treatedcount)
    
    timetest = {}
    for time in timepoints:
        test = {}
        if(len(gender)>1 and gendercount == 1):
            for sex in gender:
                if(len(genotype)>1 and genotypecount == 1):
                    #print("First flag")
                    for gene in genotype:
                        if(len(induced)>1 and inducedcount == 1):
                            for inducedlabel in induced:
                                if(len(treated)>1 and treatedcount == 1):
                                    for treatedlabel in treated:
                                        #To DO
                                        m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex,induced= inducedlabel,treated=treatedlabel)
                                        parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                                        label = sex + " "+gene +" "+inducedlabel + " " + treatedlabel
                                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                                        df.dropna()
                                        test[label] = df
                                else:
                                    m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex,induced= inducedlabel)
                                    parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                                    label = sex + " "+gene +" "+inducedlabel
                                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                                    df.dropna()
                                    test[label] = df
                        elif(len(treated)>1 and treatedcount == 1):
                            m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex,treated=treatedlabel)
                            parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                            label = sex + " "+gene +" "+treatedlabel
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                            df.dropna()
                            test[label] = df                  	
                        else:
                            m = Mouse.objects.filter(id__in= mouselist,genotype=gene,gender=sex)
                            parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                            label = sex + " "+gene
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                            df.dropna()
                            test[label] = df
                elif(len(induced)>1 and inducedcount == 1):
                    for inducedlabel in induced:
                        if(len(treated)>1 and treatedcount == 1):
                           for treatedlabel in treated:
                                m = Mouse.objects.filter(id__in= mouselist,gender=sex,induced= inducedlabel,treated=treatedlabel)
                                parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                                label = sex +" "+inducedlabel + " " + treatedlabel
                                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                                df.dropna()
                                test[label] = df
                        else:
                            m = Mouse.objects.filter(id__in= mouselist,gender=sex,induced= inducedlabel)
                            parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                            label = sex + " "+inducedlabel
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                            df.dropna()
                            test[label] = df
                elif(len(treated)>1 and treatedcount == 1):
                    m = Mouse.objects.filter(id__in= mouselist,gender=sex,treated=treatedlabel)
                    parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                    label = sex+" "+treatedlabel
                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                    df.dropna()  
                    test[label] = df

                else:
                    m = Mouse.objects.filter(id__in= mouselist,gender=sex)
                    parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                    label = sex
                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                    df.dropna()
                    test[label] = df
        elif(len(genotype)>1 and genotypecount == 1):
            for gene in genotype:
                if(len(induced)>1 and inducedcount == 1):
                    for inducedlabel in induced:
                        if(len(treated)>1 and treatedcount == 1):
                            for treatedlabel in treated:
                                m = Mouse.objects.filter(id__in= mouselist,genotype=gene,induced= inducedlabel,treated=treatedlabel)
                                parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                                label = gene +" "+inducedlabel + " " + treatedlabel
                                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                                df.dropna()
                                test[label] = df
                        else:
                            m = Mouse.objects.filter(id__in= mouselist,genotype=gene,induced= inducedlabel)
                            parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                            label = gene +" "+inducedlabel
                            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                            df.dropna()
                            test[label] = df
                elif(len(treated)>1 and treatedcount == 1):
                    m = Mouse.objects.filter(id__in= mouselist,genotype=gene,treated=treatedlabel)
                    parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                    label = gene +" "+treatedlabel
                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                    df.dropna()
                    test[label] = df
                else:
                    m = Mouse.objects.filter(id__in= mouselist,genotype=gene)
                    parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                    label = gene
                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                    df.dropna()
                    test[label] = df
        elif(len(induced)>1 and inducedcount == 1):
            for inducedlabel in induced:
                if(len(treated)>1 and treatedcount == 1):
                    for treatedlabel in treated:
                        m = Mouse.objects.filter(id__in= mouselist,induced= inducedlabel,treated=treatedlabel)
                        parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                        label = inducedlabel + " " + treatedlabel
                        df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                        df.dropna()
                        test[label] = df
                else:
                    m = Mouse.objects.filter(id__in= mouselist,induced= inducedlabel)
                    parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                    label = inducedlabel
                    df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                    df.dropna()
                    test[label] = df
        elif(len(treated)>1 and treatedcount == 1):
            for treatedlabel in treated:
                m = Mouse.objects.filter(id__in= mouselist,genotype=gene,treated=treatedlabel)
                parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                label = treatedlabel
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
        else:
            if(all==False):
                m = Mouse.objects.filter(id__in= mouselist)
                parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
                label = 'All'
                df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
                df.dropna()
                test[label] = df
        if(all==True):
            m = Mouse.objects.filter(id__in= mouselist)
            parameter_measures = measures.filter(mid__in = m,timepoint=time['timepoint']).values_list('timepoint',parameter)
            label = 'All'
            df = pd.DataFrame(list(parameter_measures.values('timepoint',parameter)))
            df.dropna()
            test[label] = df   	
        if(df.empty):
        	flag=True
        timetest[time['timepoint']] = test
        #print(test)
    #print('Out Parameter Measures')
    #clean_dict = filter(lambda k: not isnan(test[k]), test)
    return timetest,flag,genotype
def findlabels(mouselist,flag,labelist,measures):
	counter = 0
	for label in labelist:
		if(flag == 1):
			m = Mouse.objects.filter(id__in= mouselist,gender=label)
		elif(flag == 2):
			m = Mouse.objects.filter(id__in= mouselist,genotype=label)
		elif(flag == 3):
			m = Mouse.objects.filter(id__in= mouselist,induced=label)
		elif(flag == 4):
			m = Mouse.objects.filter(id__in= mouselist,treated=label)
		parameter_measures = measures.filter(mid__in = m)
		#print("Find labels")
		#print(parameter_measures)
		if(len(parameter_measures)):
			counter = counter + 1
	if(counter==2):
		return 1;
	return -1

##########################################################################################
#             PLOT   TYPES                  ##############################################
##########################################################################################

def boxplot(measures,parameter,title,xtext,ytext):
	data = {}
	data['data'] = []
	# Get parameter measurements 
	[test,flag,genotype] = parameterMeasures2(measures,parameter,False)
	symbols = ['circle', 'circle-open','square-open-dot','circle-open-dot']
	i = 0
	for key in test:
		selected_rows = test[key].dropna()
		if(not selected_rows.empty):
			test2 = test[key].sort_values('timepoint').reset_index()
			timepoint = test2['timepoint'].values.tolist()
			parametervalue = test2[parameter].values.tolist()
			parametervalue = [item for item in parametervalue if not(isnan(item)) == True]
			print("ISNAN",parametervalue)
			trace4 = {
				'y': parametervalue,
				'x': timepoint,
				'boxpoints': 'all',
				'name': key,
				'jitter': 0.5,
				'pointpos': 0.0,
				'type': 'box',
				'boxmean': 'yes',
				'marker': {
					'symbol': symbols[i],
					'size': 5,
					}
				};
			i=i+1
			data['data'].append(trace4)
	layout = {
			'title': {
				'text':title,
				'font': {
					'family': 'Source Sans Pro',
					'size': 24,
				},
			},
			'xaxis': {
				'title': xtext,
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 'false',			},
			'yaxis': {
				'title': ytext,
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 0,
			},
			'boxmode': 'group',
			'colorway': ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4'],
		}
	data['layout']=layout
	return data

def barplot(measures,parameter,title,xtext,ytext):
	data = {}
	data['data'] = []
	# Get parameter measurements 
	[test,flag,genotype] = parameterMeasures2(measures,parameter,False)
	symbols = ['circle', 'circle-open','square-open-dot','circle-open-dot']
	colorway= ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4']
	i = 0
	for key in test:
		selected_rows = test[key].dropna()
		if(not selected_rows.empty):
			test2 = test[key].groupby('timepoint').mean().reset_index()
			test3 = test[key].sort_values('timepoint').reset_index()
			timepoint = test2['timepoint'].values.tolist()
			timepoint3 = test3['timepoint'].values.tolist()
			timepoint31 = [item - 0.4 for item in timepoint3]
			timepoint32 = [item + 0.4 for item in timepoint3]
			parametervalue = test2[parameter].values.tolist()
			parametervalue = [item for item in parametervalue if not(isnan(item)) == True]
			parametervalueall = test3[parameter].values.tolist()
			parametervalueall = [item for item in parametervalueall if not(isnan(item)) == True]
			if(i==0):
				timepoint3 = timepoint31
			else:
				timepoint3 = timepoint32
			trace4 = {
				'y': parametervalue,
				'x': timepoint,
				'points': parametervalueall,
				'name': key,
				'jitter': 0.5,
				'pointpos': 0.0,
				'type': 'bar',
				'mode': 'bars+markers',
				'marker': {
					'y': parametervalueall,
					'x': timepoint3,
					'symbol': symbols[i],
					'type': 'data',
					'size': 5,
					},
				'boxmean': 'yes',
				};
			trace5 = {
				'y': parametervalueall,
				'x': timepoint3,
				'name': key,
				'jitter': 0.5,
				'type': 'scatter',
				'mode': 'bars+markers',
				'marker': {
					'symbol': symbols[i],
					'color': colorway[i],
					'size': 5,
					},
				};
			i=i+1
			data['data'].append(trace4)
			data['data'].append(trace5)
	layout = {
			'title': {
				'text':title,
				'font': {
					'family': 'Source Sans Pro',
					'size': 24,
				},
			},
			'xaxis': {
				'title': xtext,
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 'false',			},
			'yaxis': {
				'title': ytext,
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 0,
			},
			'boxmode': 'group',
			'colorway': ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4'],
		}
	data['layout']=layout
	return data

def barplot_errors(measures,parameter,title,xtext,ytext):
	data = {}
	data['data'] = []
	# Get parameter measurements 
	[test,flag,genotype] = parameterMeasures2(measures,parameter,False)
	symbols = ['circle', 'circle-open','square-open-dot','circle-open-dot']
	colorway= ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4']
	i = 0
	for key in test:
		selected_rows = test[key].dropna()
		if(not selected_rows.empty):
			test2 = test[key].groupby('timepoint').mean().reset_index()
			testmax = test[key].groupby('timepoint').max().reset_index()
			test3 = test[key].sort_values('timepoint').reset_index()
			timepoint = test2['timepoint'].values.tolist()
			timepoint3 = test3['timepoint'].values.tolist()
			parametervalue = test2[parameter].values.tolist()
			parametervaluemax = testmax[parameter].values.tolist()
			subtracted = [element1 - element2 for (element1, element2) in zip(parametervaluemax, parametervalue)] 		
			parametervalue = [item for item in parametervalue if not(isnan(item)) == True]
			parametervalueall = test3[parameter].values.tolist()
			parametervalueall = [item for item in parametervalueall if not(isnan(item)) == True]
			trace4 = {
				'y': parametervalue,
				'x': timepoint,
				'points': parametervalueall,
				'name': key,
				'error_y': {
					'type': 'data',
					'array': subtracted
				},
				'jitter': 0.5,
				'pointpos': 0.0,
				'type': 'bar',
				'mode': 'bars+markers',
				'marker': {
					'y': parametervalueall,
					'x': timepoint3,
					'symbol': symbols[i],
					'type': 'data',
					'size': 5,
					},
				'boxmean': 'yes',
				};
			i=i+1
			data['data'].append(trace4)
	layout = {
			'title': {
				'text':title,
				'font': {
					'family': 'Source Sans Pro',
					'size': 24,
				},
			},
			'xaxis': {
				'title': xtext,
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 'false',			},
			'yaxis': {
				'title': ytext,
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 0,
			},
			'boxmode': 'group',
			'colorway': ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4'],
		}
	data['layout']=layout
	return data

def barplot_errors_parameters(measures,time_var,parameters,names,title,xtext,ytext):
	data2 = {}
	data2['data'] = []
	# Get parameter measurements
	symbols = ['circle', 'circle-open','square-open-dot','circle-open-dot']
	colorway= ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4']
	i = 0

	parameter_values = {}
	timepoints = measures.values('timepoint').distinct().order_by('timepoint')
	print([time['timepoint'] for time in timepoints])
	data = {}
	for parameter in parameters:
		print("Selected parameter",parameter)
		[test,flag,genotype] = parameterMeasures4(measures,parameter,time_var)
		for key in test:
			if(key in data):				
				data[key]['parameters'].append(parameter)
				data[key]['values'].append(test[key][parameter].mean())
				data[key]['error_y'].append(test[key][parameter].max()-test[key][parameter].mean())
			else:
				data[key] = {'parameters': [], 'values': [], 'error_y': []}
				data[key]['parameters'].append(parameter)
				data[key]['values'].append(test[key][parameter].mean())
				data[key]['error_y'].append(test[key][parameter].max()-test[key][parameter].mean())
	print("Returned form parametermeasures3 ",data)
	for key in data:
		trace4 = {
			'y': data[key]['values'],
			'x': data[key]['parameters'],
			'points': data[key]['values'],
			'name': key,
			'error_y': {
				'type': 'data',
				'array': data[key]['error_y']
			},
			'jitter': 0.5,
			'pointpos': 0.0,
			'type': 'bar',
			'mode': 'bars+markers',
			'boxmean': 'yes',
			};
		i=i+1
		data2['data'].append(trace4)

	return data2

#############################################
#              NANOU                        #
#############################################

def plot_hpibd02(test,parameter):
	trace4 = {
		'y': [0, 1, 1, 2, 3, 5, 8, 13, 21],
		'boxpoints': 'all',
		'name':"test1",
		'jitter': 0.5,
		'pointpos': 0.0,
		'type': 'box',
		'boxmean': '1',
		'marker': {
			'symbol': 'circle',
			'size': 8,
			'color':'black'
		},
	};
	trace2= {
		'y': [0, 1, 5, 2, 3, 5, 9, 13, 21],
		'name':"test2",
		'boxpoints': 'all',
		'jitter': 1.5,
		'pointpos': 0.0,
		'type': 'box',
		'boxmean': "true",
		'marker': {
			'symbol': 'circle-open-dot',
			'size': 8,
			'color':'black'
		}
	};

	yaxis= {
		'title':'USD (millions)'
	};
	layout = {
			'title': {
				'text':'IINFLC-04: Body Weight',
				'font': {
					'family': 'Source Sans Pro',
					'size': 24,
				},
			},
			'xaxis': {
				'title':'Days post immunization',
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 'false',			},
			'yaxis': {
				'title':'Weight (%)',
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'normal',
				'showgrid': 0,
				'range': [50, 120]
			},
			'colorway': ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4'],
		}
	data = [trace4,trace2]

	return [data,layout]

def plot_iinflc04(measures):
	parameter = "weight"
	data2 = {}
	[test,flag,genotype] = parameterMeasures2(measures,"weight",False)
	data = []
	symbols = ['square', 'circle','square-open-dot','circle-open-dot']
	i = 0
	data2[parameter]={}
	data2[parameter]['data'] = []
	#data2[parameter]['layout'] = {}
	for key in test:
		selected_rows = test[key].dropna()
		if(not selected_rows.empty):
			test2 = test[key].groupby('timepoint').mean().reset_index()
			testmin = test[key].groupby('timepoint').min().reset_index()
			testmax= test[key].groupby('timepoint').max().reset_index()
			subtracted = [element1 - element2 for (element1, element2) in zip(testmax[parameter].values.tolist(), test2[parameter].values.tolist())]
			timepoint = test2['timepoint'].values.tolist()
			parametervalue = test2[parameter].values.tolist()
			parametervalue = [x*100 / parametervalue[0] for x in parametervalue]
			print('Mean',parametervalue)
			print('Error',subtracted)
			print(key)
			trace1= {
				'x': timepoint,
				'y': parametervalue,
				'error_y': {
					'type': 'data',
					'array': subtracted,
				},
				'name': key,
				'mode': 'lines+markers',
				'hovertemplate': '<i>Day</i>: %{x}' +

		                        '<br><b>Weight(%)</b>: %{y}<br>',
		        'marker': {
		        	'symbol': symbols[i],
		        	'size': 8
				},
				'boxpoints': 'all',
				'jitter': 1.5,
				'pointpos': 0.0,
				'type': 'scatter',
				'boxmean': '1',
			};
			i=i+1
			data.append(trace1)
			data2[parameter]['data'].append(trace1)

	layout = {
			'title': {
				'text':'IINFLC-04: Body Weight',
				'font': {
					'family': 'Source Sans Pro',
					'size': 24,
				},
			},
			'xaxis': {
				'title':'Days post immunization',
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 'false',			},
			'yaxis': {
				'title':'Weight (%)',
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'normal',
				'showgrid': 0,
				'range': [50, 120]
			},
			'colorway': ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4'],
		}

	data2[parameter]['layout'] = layout
	print(data2[parameter]['data'])
	return data2

def plot_ni01(measures):
	parameter = "clinical_score"
	data2 = {}
	data2[parameter]={}
	data2[parameter]['data'] = []
	[test,flag,genotype] = parameterMeasures2(measures,"clinical_score",False)
	data = []
	symbols = ['square', 'circle','square-open-dot','circle-open-dot']
	i = 0

	for key in test:
		selected_rows = test[key].dropna()
		if(not selected_rows.empty):
			test2 = test[key].groupby('timepoint').mean().reset_index()
			testmin = test[key].groupby('timepoint').min().reset_index()
			testmax= test[key].groupby('timepoint').max().reset_index()
			subtracted = [element1 - element2 for (element1, element2) in zip(testmax[parameter].values.tolist(), test2[parameter].values.tolist())]
			timepoint = test2['timepoint'].values.tolist()
			parametervalue = test2[parameter].values.tolist()
			subtracted2 = [element1 - element2 for (element1, element2) in zip(testmax[parameter].values.tolist(), parametervalue)]
			error = [(element1 + element2)/2 for (element1, element2) in zip(subtracted,subtracted2)]
			trace1= {
				'x': timepoint,
				'y': parametervalue,
				'error_y': {
					'type': 'data',
					'array': error,
				},
				'name': key,
				'mode': 'lines+markers',
				'hovertemplate': '<i>Day</i>: %{x}' +

		                        '<br><b>Clinical score</b>: %{y}<br>',
		        'marker': {
		        	'symbol': symbols[i],
		        	'size': 8
				},
				'boxpoints': 'all',
				'jitter': 1.5,
				'pointpos': 0.0,
				'type': 'scatter',
				'boxmean': '1',
			};
			i=i+1
			data.append(trace1)
			data2[parameter]['data'].append(trace1)

	layout = {
			'title': {
				'text':'NI-01: Clinical score',
				'font': {
					'family': 'Source Sans Pro',
					'size': 24,
				},
			},
			'xaxis': {
				'title':'Days post immunization',
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 'false',			},
			'yaxis': {
				'title':'Clinical score',
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 0,
				'range':[0,5]
			},
			'colorway': ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4'],
		}
	data2[parameter]['layout']= layout
	return data2

def plot_ni02grs01(measures):
	data2 = {}
	data2['plot1'] = {}
	parameter = "forelimb_mean_ratio"
	data2['plot1'] = boxplot(measures,parameter,'NI-02-GRS-01: Grip strength (all)','Days post immunization','Grip strength (g)')
	data2['plot2'] = barplot(measures,parameter,'NI-02-GRS-01: Grip strength (all)','Days post immunization','Grip strength (g)')
	return data2

def plot_hem01(measures):
	data2 = {}
	data2['rbc'] = {}
	parameter = "rbc_count"
	data2['rbc'] = barplot_errors(measures,"rbc_count",'HEM-01: RBC','Days post immunization','RBC (10^12/L)')
	data2['hbg'] = barplot_errors(measures,"hb",'HEM-01: HGB','Days post immunization','HGB (g/dL)')
	data2['hct'] = barplot_errors(measures,"ht",'HEM-01: HCT','Days post immunization','HCT (%)')
	data2['mcv'] = barplot_errors(measures,"mcv",'HEM-01: MCV','Days post immunization','MCV (fL)')
	data2['mch'] = barplot_errors(measures,"mch",'HEM-01: MCH','Days post immunization','MCH (pg)')
	data2['mchc'] = barplot_errors(measures,"mchc",'HEM-01: MCHC','Days post immunization','MCHC (g/dL)')
	data2['rdwcv'] = barplot_errors(measures,"rdwcv",'HEM-01: RDW-CV','Days post immunization','RDW-CV (%)')
	data2['rdwsd'] = barplot_errors(measures,"rdwsd",'HEM-01: RDW-SD','Days post immunization','RDW-SD (fL)')
	data2['plt'] = barplot_errors(measures,"plt",'HEM-01: PLT','Days post immunization','PLT')
	data2['mpv'] = barplot_errors(measures,"mpv",'HEM-01: MPV','Days post immunization','MPV (fL)')
	data2['pdw'] = barplot_errors(measures,"pdw",'HEM-01: PDW','Days post immunization','PDW')
	data2['pct'] = barplot_errors(measures,"pct",'HEM-01: PCT','Days post immunization','PCT (%)')

	parameters = ['wbc_count','neutrophils_num','lymphocytes_num','eosinophils_num','basophils_num','monocytes_num']
	names = ['WBC','Neu #','Lym #','Mon #','Eos #','Bas #']

	timepoints = measures.values('timepoint').distinct().order_by('timepoint')
	for time in timepoints:
		data2[time['timepoint']] = {}
	for time in timepoints:
		
		data2[time['timepoint']] = barplot_errors_parameters(measures,time['timepoint'],parameters, names,'HEM-01: Blood cell counts ('+str(time['timepoint']) +' dpi)',' ','# Cells')
		print("-------------------"+ str(time['timepoint']) +"---------------------------")
		print(data2[time['timepoint']])
		print("-------------------------------------------------------------------")
	
	print("return to views", data2)
	return data2,timepoints
#############################################
#              KOLIARAKI                   #
#############################################
def plot_iinflc01(test,parameter):
	data = []
	symbols = ['square', 'circle','square-open-dot','circle-open-dot']
	i = 0
	for key in test:
		if(len(test[key])):
			test2 = test[key].groupby('timepoint').mean().reset_index()
			if(len(test2)):
				print("Inside iinflc01", test2)
				testmin = test[key].groupby('timepoint').min().reset_index()
				testmax= test[key].groupby('timepoint').max().reset_index()
				subtracted = [element1 - element2 for (element1, element2) in zip(testmax[parameter].values.tolist(), test2[parameter].values.tolist())]
				timepoint = test2['timepoint'].values.tolist()
				parametervalue = test2[parameter].values.tolist()
				parametervalue = [x*100 / parametervalue[0] for x in parametervalue]
				print('Mean',parametervalue)
				print('Error',subtracted)
				print(key)
				trace1= {
					'x': timepoint,
					'y': parametervalue,
					'error_y': {
						'type': 'data',
						'array': subtracted,
					},
					'name': key,
					'mode': 'lines+markers',
					'hovertemplate': '<i>Day</i>: %{x}' +

			                        '<br><b>Weight(%)</b>: %{y}<br>',
			        'marker': {
			        	'symbol': symbols[i],
			        	'size': 8
					},
					'boxpoints': 'all',
					'jitter': 1.5,
					'pointpos': 0.0,
					'type': 'scatter',
					'boxmean': '1',
				};
				i=i+1
				data.append(trace1)

	layout = {
			'title': {
				'text':'IINFLC-01: Body Weight',
				'font': {
					'family': 'Source Sans Pro',
					'size': 24,
				},
			},
			'xaxis': {
				'title':'Days of treatement',
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'tozero',
				'showgrid': 'false',			},
			'yaxis': {
				'title':'Weight (%)',
				'font': {
					'family': 'Source Sans Pro',
					'size': 14,
				},
				'rangemode': 'normal',
				'showgrid': 0,
			},
			'colorway': ['#969c9c', '#4ed2c5', '#0c8b94', '#303144', '#212529', '#FF9655', '#FFF263', '#6AF9C4'],
		}
	return [data,layout]
